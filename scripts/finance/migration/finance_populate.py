#!/usr/bin/env python3
"""Popula Categorias (65 rows do Excel) e Contas (5 origens) no Notion."""
import json
import os
import sys
import warnings
from pathlib import Path

import requests
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

TOKEN = os.environ["NOTION_TOKEN"]
CAT_DB = os.environ["NOTION_FH_CATEGORIAS"]
CTA_DB = os.environ["NOTION_FH_CONTAS"]
HEADERS = {
    "Authorization": f"Bearer {TOKEN}",
    "Notion-Version": "2022-06-28",
    "Content-Type": "application/json",
}
EXCEL = "/Users/enzopruano/Desktop/Contas/Contas Enzo-Ana.xlsx"

# ---------------- Categorias ----------------
wb = load_workbook(EXCEL, data_only=True)
s = wb["Categorias"]
cat_rows = []
for row in s.iter_rows(min_row=2, values_only=True):
    macro, sub, icon = (row[0], row[1], row[2] if len(row) > 2 else "")
    if not macro or not sub:
        continue
    cat_rows.append({"macro": macro.strip(), "sub": sub.strip(), "icon": (icon or "").strip()})

def tipo_for(macro: str) -> str:
    if macro == "Receita":
        return "Receita"
    if macro == "Transferências":
        return "Transferência"
    return "Despesa"

cat_id_map = {}  # "Macro > Sub" → notion page id

print(f"Populando {len(cat_rows)} categorias…")
for c in cat_rows:
    nome = f"{c['macro']} > {c['sub']}"
    body = {
        "parent": {"database_id": CAT_DB},
        "icon": {"type": "emoji", "emoji": c["icon"] or "📦"} if c["icon"] else None,
        "properties": {
            "Nome": {"title": [{"text": {"content": nome}}]},
            "Macro": {"select": {"name": c["macro"]}},
            "Subcategoria": {"rich_text": [{"text": {"content": c["sub"]}}]},
            "Ícone": {"rich_text": [{"text": {"content": c["icon"]}}]} if c["icon"] else {"rich_text": []},
            "Tipo": {"select": {"name": tipo_for(c["macro"])}},
        },
    }
    if body["icon"] is None:
        del body["icon"]
    r = requests.post("https://api.notion.com/v1/pages", headers=HEADERS, json=body, timeout=30)
    if r.status_code >= 300:
        print(f"ERRO '{nome}': {r.text[:300]}", file=sys.stderr)
        sys.exit(1)
    cat_id_map[nome] = r.json()["id"]
    print(f"  {c['icon']} {nome}")

# ---------------- Contas ----------------
contas = [
    {"nome": "BTG Conta",     "tipo": "Conta Corrente",    "banco": "BTG",           "emoji": "🏦"},
    {"nome": "BTG Cartão",    "tipo": "Cartão de Crédito", "banco": "BTG",           "emoji": "💳"},
    {"nome": "Itaú Conta",    "tipo": "Conta Corrente",    "banco": "Itaú",          "emoji": "🏦"},
    {"nome": "Itaú Cartão",   "tipo": "Cartão de Crédito", "banco": "Itaú",          "emoji": "💳"},
    {"nome": "Mercado Pago",  "tipo": "Carteira Digital",  "banco": "Mercado Pago",  "emoji": "💼"},
]
cta_id_map = {}
print(f"\nPopulando {len(contas)} contas…")
for c in contas:
    body = {
        "parent": {"database_id": CTA_DB},
        "icon": {"type": "emoji", "emoji": c["emoji"]},
        "properties": {
            "Nome": {"title": [{"text": {"content": c["nome"]}}]},
            "Tipo": {"select": {"name": c["tipo"]}},
            "Banco": {"select": {"name": c["banco"]}},
            "Ativa": {"checkbox": True},
        },
    }
    r = requests.post("https://api.notion.com/v1/pages", headers=HEADERS, json=body, timeout=30)
    if r.status_code >= 300:
        print(f"ERRO '{c['nome']}': {r.text[:300]}", file=sys.stderr)
        sys.exit(1)
    cta_id_map[c["nome"]] = r.json()["id"]
    print(f"  {c['emoji']} {c['nome']}")

# ---------------- salva maps ----------------
maps_dir = Path(os.environ["HOME"]) / "Agent"
(maps_dir / ".notion-finance-maps.json").write_text(json.dumps({
    "categorias": cat_id_map,
    "contas": cta_id_map,
}, indent=2, ensure_ascii=False))
print(f"\n✅ {len(cat_rows)} categorias + {len(contas)} contas populadas.")
print(f"Maps salvos em ~/Agent/.notion-finance-maps.json")
