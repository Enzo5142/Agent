#!/usr/bin/env python3
"""Importa 821 transações do Excel → Notion Transações DB."""
import json
import os
import sys
import time
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

import requests
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

TOKEN = os.environ["NOTION_TOKEN"]
TX_DB = os.environ["NOTION_FH_TRANSACOES"]
HEADERS = {
    "Authorization": f"Bearer {TOKEN}",
    "Notion-Version": "2022-06-28",
    "Content-Type": "application/json",
}
EXCEL = "/Users/enzopruano/Desktop/Contas/Contas Enzo-Ana.xlsx"

maps = json.loads(open(os.path.expanduser("~/Agent/.notion-finance-maps.json")).read())
CAT_MAP = maps["categorias"]  # "Moradia > Aluguel" → id
CTA_MAP = maps["contas"]       # "BTG Conta" → id


def parse_date(s):
    if isinstance(s, datetime):
        return s.date().isoformat()
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(s).strip(), fmt).date().isoformat()
        except ValueError:
            continue
    return None


def build_body(row):
    data, desc, cat, sub, tipo, valor, forma, origem, parc, fixa, mesref, obs = row[:12]
    cat_key = f"{cat} > {sub}"
    cat_id = CAT_MAP.get(cat_key)
    cta_id = CTA_MAP.get(origem)
    iso = parse_date(data)
    props = {
        "Descrição": {"title": [{"text": {"content": str(desc)[:2000]}}]},
        "Data": {"date": {"start": iso}} if iso else {"date": None},
        "Tipo": {"select": {"name": tipo}} if tipo else {"select": None},
        "Valor": {"number": float(valor) if valor is not None else None},
        "MesRef": {"rich_text": [{"text": {"content": str(mesref)}}]} if mesref else {"rich_text": []},
    }
    if cat_id:
        props["Categoria"] = {"relation": [{"id": cat_id}]}
    if cta_id:
        props["Conta"] = {"relation": [{"id": cta_id}]}
    if forma:
        props["Forma Pgto"] = {"select": {"name": forma}}
    if parc:
        props["Parcela"] = {"rich_text": [{"text": {"content": str(parc)}}]}
    if fixa:
        props["Fixa/Variável"] = {"select": {"name": fixa}}
    if obs:
        props["Observações"] = {"rich_text": [{"text": {"content": str(obs)[:2000]}}]}
    return {"parent": {"database_id": TX_DB}, "properties": props}


def post_row(idx, row):
    body = build_body(row)
    for attempt in range(5):
        r = requests.post("https://api.notion.com/v1/pages", headers=HEADERS, json=body, timeout=30)
        if r.status_code == 200:
            return (idx, True, None)
        if r.status_code == 429:
            time.sleep(1 + attempt)
            continue
        if r.status_code >= 500:
            time.sleep(0.5 + attempt * 0.5)
            continue
        return (idx, False, r.text[:400])
    return (idx, False, "max retries")


# ---------------- main ----------------
wb = load_workbook(EXCEL, data_only=True)
rows = [r for r in wb["Transações"].iter_rows(min_row=2, values_only=True) if any(r)]
print(f"Carregadas {len(rows)} transações do Excel.")

# check coverage
missing_cat = {f"{r[2]} > {r[3]}" for r in rows if f"{r[2]} > {r[3]}" not in CAT_MAP}
missing_cta = {r[7] for r in rows if r[7] not in CTA_MAP}
if missing_cat:
    print(f"⚠️  Categorias não mapeadas: {missing_cat}", file=sys.stderr)
if missing_cta:
    print(f"⚠️  Contas não mapeadas: {missing_cta}", file=sys.stderr)
if missing_cat or missing_cta:
    print("Abortando — mapear primeiro.", file=sys.stderr)
    sys.exit(1)

ok = 0
fail = []
t0 = time.time()
# 4 threads respeita limite de ~3 req/s em média com latency de rede
with ThreadPoolExecutor(max_workers=4) as ex:
    futures = [ex.submit(post_row, i, r) for i, r in enumerate(rows)]
    for f in as_completed(futures):
        idx, success, err = f.result()
        if success:
            ok += 1
            if ok % 50 == 0:
                print(f"  {ok}/{len(rows)}…")
        else:
            fail.append((idx, err))

elapsed = time.time() - t0
print(f"\n✅ {ok} transações importadas em {elapsed:.1f}s.")
if fail:
    print(f"❌ {len(fail)} falhas:")
    for i, err in fail[:5]:
        print(f"  row {i}: {err}")
