"""Parser da fatura de cartão BTG.

Formato: .xlsx criptografado (Agile Encryption AES-128-CBC SHA1).
Senha = CPF do Enzo (507.840.248-98 sem pontos = 50784024898).
Sheet: 'Titular'.
Colunas: B(data), C(desc), E(valor), F(tipo), H(cartão).

MesRef = mês do vencimento extraído do título da fatura
(ex: 'Janeiro/2026', 'Fevereiro/2026').
"""
from __future__ import annotations

import io
import re
from datetime import date, datetime
from pathlib import Path

import msoffcrypto
from openpyxl import load_workbook

from rules import RawTransaction


CPF_SENHA = "50784024898"

MESES_PT = {
    "janeiro": 1, "fevereiro": 2, "março": 3, "marco": 3, "abril": 4,
    "maio": 5, "junho": 6, "julho": 7, "agosto": 8, "setembro": 9,
    "outubro": 10, "novembro": 11, "dezembro": 12,
}


def _decrypt(path: Path) -> io.BytesIO:
    buf = io.BytesIO()
    with open(path, "rb") as f:
        office = msoffcrypto.OfficeFile(f)
        office.load_key(password=CPF_SENHA)
        office.decrypt(buf)
    buf.seek(0)
    return buf


def detect(path: Path) -> bool:
    if path.suffix.lower() != ".xlsx":
        return False
    # Assinatura OLE2 (xlsx criptografado é OLE2 container)
    with open(path, "rb") as f:
        head = f.read(8)
    if head[:4] != b"\xd0\xcf\x11\xe0":
        return False
    # Tenta decriptar com a senha do Enzo
    try:
        decrypted = _decrypt(path)
        wb = load_workbook(decrypted, read_only=True)
        return "Titular" in wb.sheetnames
    except Exception:
        return False


def _extract_vencimento(wb) -> date | None:
    """Procura nas primeiras 10 linhas de todas as sheets algo tipo
    'Fatura Janeiro/2026' ou 'Vencimento DD/MM/YYYY'.
    """
    for sheet in wb.sheetnames:
        s = wb[sheet]
        rows_read = 0
        for row in s.iter_rows(values_only=True):
            if rows_read >= 15:
                break
            rows_read += 1
            for cell in row:
                if cell is None:
                    continue
                text = str(cell).strip()
                # Tenta "Mês/Ano"
                m = re.search(r"(janeiro|fevereiro|mar[çc]o|abril|maio|junho|julho|agosto|setembro|outubro|novembro|dezembro)\s*/\s*(\d{4})", text, re.I)
                if m:
                    mes = MESES_PT[m.group(1).lower()]
                    ano = int(m.group(2))
                    # Vencimento fictício dia 25 (data exata menos relevante, mês importa)
                    return date(ano, mes, 25)
                # Tenta "DD/MM/YYYY" precedido por Vencimento
                m2 = re.search(r"vencimento[^\d]*(\d{2})/(\d{2})/(\d{4})", text, re.I)
                if m2:
                    return date(int(m2.group(3)), int(m2.group(2)), int(m2.group(1)))
    return None


def _parse_date(raw) -> date | None:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, str):
        for fmt in ("%d/%m/%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(raw.strip(), fmt).date()
            except ValueError:
                continue
    return None


def parse(path: Path) -> list[RawTransaction]:
    decrypted = _decrypt(path)
    wb = load_workbook(decrypted, read_only=True, data_only=True)

    venc = _extract_vencimento(wb)
    if venc is None:
        raise RuntimeError(f"Não achei vencimento na fatura BTG Cartão {path.name}")

    s = wb["Titular"]
    # Acha linha de header
    header_row = None
    for idx, row in enumerate(s.iter_rows(values_only=True), 1):
        row_text = " ".join(str(c or "").lower() for c in row)
        if ("data" in row_text and "valor" in row_text) or "descri" in row_text:
            header_row = idx
            break
        if idx > 15:
            break
    if header_row is None:
        header_row = 1

    txs: list[RawTransaction] = []
    for idx, row in enumerate(s.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        # B=1, C=2, E=4, F=5, H=7 (0-indexed)
        if len(row) < 8:
            continue
        data_raw = row[1]
        desc = str(row[2] or "").strip()
        val_raw = row[4]
        tipo_str = str(row[5] or "").strip()
        cartao = str(row[7] or "").strip()

        if not desc:
            continue

        data = _parse_date(data_raw)
        if data is None:
            continue
        try:
            valor = float(val_raw)
        except (TypeError, ValueError):
            continue
        if valor == 0:
            continue

        parcela = None
        m = re.search(r"\b(\d{1,2})/(\d{1,2})\b", desc)
        if m:
            parcela = f"{int(m.group(1)):02d}/{int(m.group(2)):02d}"

        txs.append(RawTransaction(
            data=data,
            descricao=desc,
            valor=abs(valor),
            origem="BTG Cartão",
            forma_pgto="Crédito",
            parcela=parcela,
            vencimento=venc,
            tipo_hint="Despesa" if valor > 0 else "Receita",  # estorno é valor negativo
            observacao=f"{cartao} | {tipo_str}" if cartao or tipo_str else None,
            raw_source=f"{path.name}:r{idx}",
        ))
    return txs
